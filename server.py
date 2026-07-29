# -*- coding: utf-8 -*-
import http.server
import socketserver
import json
import os
import sys
import urllib.parse
import csv
import datetime
import sqlite3
import hashlib
import secrets
import shutil
import db

# Force UTF-8 for Windows embedded Python
if hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except Exception:
        pass

PORT = 8000
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def parse_12h_to_24h_minutes(t):
    """Parse '08:30 AM' or '04:00 PM' to minutes since midnight (24h internal)."""
    t = t.strip()
    parts = t.split()
    time_parts = parts[0].split(':')
    h, m = int(time_parts[0]), int(time_parts[1])
    ampm = parts[1].upper() if len(parts) > 1 else ''
    if ampm == 'PM' and h != 12:
        h += 12
    elif ampm == 'AM' and h == 12:
        h = 0
    return h * 60 + m

def parse_24h_to_minutes(t):
    """Parse '08:30' (24h string) to minutes since midnight."""
    t = t.strip()
    parts = t.split(':')
    return int(parts[0]) * 60 + int(parts[1])

def minutes_to_hhmm(minutes):
    """Convert minutes since midnight to 'HH:MM' 24h string."""
    h = (minutes // 60) % 24
    m = minutes % 60
    return f"{h:02d}:{m:02d}"

def minutes_to_12h(minutes):
    """Convert minutes since midnight to 'hh:MM AM/PM' display string."""
    h = (minutes // 60) % 24
    m = minutes % 60
    ampm = 'AM' if h < 12 else 'PM'
    h12 = h % 12 or 12
    return f"{h12:02d}:{m:02d} {ampm}"

def get_employee_shift_times(emp):
    """Get configured shift start/end in minutes for an employee."""
    if emp['shift_start_time'] and emp['shift_end_time']:
        return parse_24h_to_minutes(emp['shift_start_time']), parse_24h_to_minutes(emp['shift_end_time'])
    shift_id = emp['default_shift']
    if shift_id and shift_id.isdigit():
        shift = db.query_db("SELECT start_time, end_time FROM shifts WHERE id=?", (int(shift_id),), one=True)
        if shift:
            return parse_24h_to_minutes(shift['start_time']), parse_24h_to_minutes(shift['end_time'])
    return parse_24h_to_minutes('08:00'), parse_24h_to_minutes('16:00')
PUBLIC_DIR = os.path.join(BASE_DIR, 'public')
LOG_FILE = os.path.join(BASE_DIR, 'bvc_server.log')

def server_log(msg):
    try:
        with open(LOG_FILE, 'a', encoding='utf-8') as f:
            f.write(datetime.datetime.now().strftime('%H:%M:%S') + ' | ' + str(msg) + '\n')
    except Exception:
        pass

class BVCRequestHandler(http.server.BaseHTTPRequestHandler):
    def end_headers(self):
        # Add CORS headers for local development
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS, PUT, DELETE')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        super().end_headers()

    def do_OPTIONS(self):
        self.send_response(204)
        self.end_headers()

    def do_GET(self):
        parsed_url = urllib.parse.urlparse(self.path)
        path = urllib.parse.unquote(parsed_url.path)
        query = urllib.parse.parse_qs(parsed_url.query)

        # Route API requests
        if path.startswith('/api/'):
            self.handle_api_get(path, query)
        else:
            self.handle_static_serve(path)

    def do_POST(self):
        parsed_url = urllib.parse.urlparse(self.path)
        path = urllib.parse.unquote(parsed_url.path)
        
        if path.startswith('/api/'):
            if path == '/api/backup/upload':
                self.upload_backup()
            elif path == '/api/inventory/import-xlsx':
                self.import_inventory_xlsx()
            else:
                self.handle_api_post(path)
        else:
            self.send_error(404, "Not Found")

    # Static File Server
    def handle_static_serve(self, path):
        if path == '/' or path == '':
            path = '/index.html'
        
        # Prevent directory traversal attacks
        safe_path = os.path.normpath(path.lstrip('/'))
        file_path = os.path.join(PUBLIC_DIR, safe_path)
        
        if not file_path.startswith(PUBLIC_DIR):
            self.send_error(403, "Forbidden")
            return

        if not os.path.exists(file_path) or os.path.isdir(file_path):
            # Fallback to index.html for SPA routing
            file_path = os.path.join(PUBLIC_DIR, 'index.html')
            if not os.path.exists(file_path):
                self.send_error(404, "File Not Found")
                return

        # Determine MIME type
        content_type = 'text/plain; charset=utf-8'
        if file_path.endswith('.html'):
            content_type = 'text/html; charset=utf-8'
        elif file_path.endswith('.css'):
            content_type = 'text/css; charset=utf-8'
        elif file_path.endswith('.js'):
            content_type = 'application/javascript; charset=utf-8'
        elif file_path.endswith('.json'):
            content_type = 'application/json; charset=utf-8'
        elif file_path.endswith('.png'):
            content_type = 'image/png'
        elif file_path.endswith('.jpg') or file_path.endswith('.jpeg'):
            content_type = 'image/jpeg'
        elif file_path.endswith('.svg'):
            content_type = 'image/svg+xml; charset=utf-8'
        elif file_path.endswith('.ico'):
            content_type = 'image/x-icon'

        try:
            with open(file_path, 'rb') as f:
                content = f.read()
            self.send_response(200)
            self.send_header('Content-Type', content_type)
            self.send_header('Content-Length', str(len(content)))
            self.send_header('Cache-Control', 'no-cache, no-store, must-revalidate')
            self.send_header('Pragma', 'no-cache')
            self.send_header('Expires', '0')
            self.end_headers()
            self.wfile.write(content)
        except Exception as e:
            self.send_error(500, f"Internal Server Error: {str(e)}")

    # API GET Router
    def handle_api_get(self, path, query):
        try:
            # Routes that need query parameters
            if path == '/api/employees':
                self.get_employees(query.get('status', [None])[0])
            elif path == '/api/attendance':
                self.get_attendance(query.get('start_date', [None])[0], query.get('end_date', [None])[0])
            elif path == '/api/transactions':
                self.get_transactions(query.get('employee_id', [None])[0])
            elif path == '/api/payroll/calculate':
                self.calculate_payroll(query.get('start_date', [None])[0], query.get('end_date', [None])[0])
            elif path == '/api/payroll/export':
                self.export_payroll_csv(query.get('start_date', [None])[0], query.get('end_date', [None])[0])
            elif path == '/api/treasury':
                self.get_treasury(query.get('type', [None])[0], query.get('date_from', [None])[0], query.get('date_to', [None])[0])
            elif path == '/api/employee-ledger':
                self.get_employee_ledger(query.get('employee_id', [None])[0])
            elif path == '/api/backup/download':
                self.download_backup(query.get('filename', [None])[0])
            elif path == '/api/users/permissions':
                self.get_user_permissions(query)
            elif path == '/api/inventory':
                if query.get('search', [None])[0] or query.get('limit', [None])[0] or query.get('offset', [None])[0] or query.get('stock', [None])[0]:
                    self.get_inventory_search(query)
                else:
                    self.get_inventory()
            elif path == '/api/search':
                self.global_search(query)
            elif path == '/api/inventory/barcode-lookup':
                self.barcode_lookup(query)
            elif path == '/api/receipt/attendance':
                self.receipt_attendance(query)
            elif path == '/api/receipt/payroll':
                self.receipt_payroll(query)
            else:
                # Simple routes (no query params needed)
                simple_routes = {
                    '/api/dashboard/summary': self.get_dashboard_summary,
                    '/api/treasury/balance': self.get_treasury_balance,
                    '/api/inventory/export-xlsx': self.export_inventory_xlsx,
                    '/api/factory-expenses': self.get_factory_expenses,
                    '/api/settings': self.get_settings,
                    '/api/users': self.get_users,
                    '/api/backup/list': self.list_backups,
                    '/api/shifts': self.get_shifts,
                }
                handler = simple_routes.get(path)
                if handler:
                    handler()
                else:
                    self.send_json_response({"error": f"Endpoint not found: {path}"}, 404)
        except Exception as e:
            server_log(f"API GET Error [{path}]: {e}")
            self.send_json_response({"error": "Internal server error"}, 500)

    # API POST Router
    def handle_api_post(self, path):
        try:
            # Routes that don't need request body
            no_body_routes = {
                '/api/backup/create': self.create_backup,
            }
            if path in no_body_routes:
                no_body_routes[path]()
                return

            data = self.get_post_data()
            body_routes = {
                '/api/employees/add': self.add_employee,
                '/api/employees/update': self.update_employee,
                '/api/employees/delete': self.delete_employee,
                '/api/employees/restore': self.restore_employee,
                '/api/attendance/log': self.log_attendance,
                '/api/attendance/fast-check': self.fast_check,
                '/api/attendance/delete': self.delete_attendance,
                '/api/transactions/add': self.add_transaction,
                '/api/transactions/delete': self.delete_transaction,
                '/api/inventory/add': self.add_inventory_item,
                '/api/inventory/update': self.update_inventory_item,
                '/api/inventory/delete': self.delete_inventory_item,
                '/api/payroll/save': self.save_payroll,
                '/api/treasury/deposit': self.add_treasury_deposit,
                '/api/treasury/withdraw': self.add_treasury_withdrawal,
                '/api/employee-ledger/add': self.add_employee_ledger_entry,
                '/api/factory-expenses/add': self.add_factory_expense,
                '/api/auth/login': self.login_user,
                '/api/users/add': self.add_user,
                '/api/users/update': self.update_user,
                '/api/users/delete': self.delete_user,
                '/api/users/reset-password': self.reset_user_password,
                '/api/users/permissions': self.save_user_permissions,
                '/api/settings/save': self.save_settings,
                '/api/backup/restore': self.restore_backup,
                '/api/backup/delete': self.delete_backup,
                '/api/inventory/import-process': self.import_process_batch,
                '/api/inventory/barcode-update': self.barcode_update,
                '/api/shifts/save': self.save_shifts,
                '/api/shifts/delete': self.delete_shift,
                '/api/export': self.export_data,
            }
            handler = body_routes.get(path)
            if handler:
                handler(data)
            else:
                self.send_json_response({"error": f"Endpoint not found: {path}"}, 404)
        except Exception as e:
            server_log(f"API POST Error [{path}]: {e}")
            self.send_json_response({"error": f"Internal server error: {str(e)}"}, 500)

    # JSON Helper Functions
    def send_json_response(self, data, status_code=200):
        try:
            response_bytes = json.dumps(data, ensure_ascii=False).encode('utf-8')
            self.send_response(status_code)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.send_header('Content-Length', str(len(response_bytes)))
            self.send_header('Cache-Control', 'no-cache, no-store, must-revalidate')
            self.send_header('Pragma', 'no-cache')
            self.send_header('Expires', '0')
            self.end_headers()
            self.wfile.write(response_bytes)
        except Exception as e:
            # Fallback if connection dropped
            pass

    def get_post_data(self):
        content_length = int(self.headers.get('Content-Length', 0))
        if content_length == 0:
            return {}
        body = self.rfile.read(content_length)
        try:
            return json.loads(body.decode('utf-8'))
        except (UnicodeDecodeError, ValueError) as e:
            server_log(f"JSON parse error: {e}")
            return {}

    def parse_multipart(self):
        """Parse multipart/form-data for file upload. Returns dict with 'filename', 'filedata', 'restore'."""
        content_type = self.headers.get('Content-Type', '')
        if 'multipart/form-data' not in content_type:
            return None
        boundary = None
        for part in content_type.split(';'):
            part = part.strip()
            if part.startswith('boundary='):
                boundary = part[len('boundary='):]
                break
        if not boundary:
            return None
        content_length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(content_length)
        boundary_bytes = ('--' + boundary).encode('utf-8')
        parts = body.split(boundary_bytes)
        result = {'filename': '', 'filedata': None, 'restore': 'false'}
        for part in parts:
            if len(part) < 10:
                continue
            if part[:2] == b'\r\n':
                part = part[2:]
            if part.endswith(b'\r\n'):
                part = part[:-2]
            header_end = part.find(b'\r\n\r\n')
            if header_end < 0:
                continue
            headers_raw = part[:header_end].decode('utf-8', errors='replace')
            file_data = part[header_end + 4:]
            if 'filename="' in headers_raw:
                fn_start = headers_raw.find('filename="') + len('filename="')
                fn_end = headers_raw.find('"', fn_start)
                result['filename'] = headers_raw[fn_start:fn_end]
                result['filedata'] = file_data
            elif 'name="restore"' in headers_raw:
                result['restore'] = file_data.decode('utf-8', errors='replace').strip()
        return result

    # API Handlers: Dashboard
    def get_dashboard_summary(self):
        emp_count = db.query_db("SELECT COUNT(*) as count FROM employees WHERE status='active'", one=True)['count']
        low_stock_count = db.query_db("SELECT COUNT(*) as count FROM inventory WHERE quantity <= min_stock", one=True)['count']

        # Treasury balance (single query)
        r = db.query_db("SELECT COALESCE(SUM(CASE WHEN type='deposit' THEN amount ELSE 0 END), 0) as deposits, COALESCE(SUM(CASE WHEN type='withdrawal' THEN amount ELSE 0 END), 0) as withdrawals FROM cashbox", one=True)
        treasury_balance = r['deposits'] - r['withdrawals']

        # Employee ledger balance
        ledger_balance = db.query_db("SELECT COALESCE(SUM(amount), 0) as s FROM employee_ledger", one=True)['s']

        # Today's attendance count
        today = datetime.date.today().isoformat()
        attendance_today = db.query_db("SELECT COUNT(*) as count FROM attendance WHERE date=?", (today,), one=True)['count']

        self.send_json_response({
            "employees_count": emp_count,
            "low_stock_count": low_stock_count,
            "treasury_balance": treasury_balance,
            "ledger_balance": ledger_balance,
            "attendance_today": attendance_today
        })

    # API Handlers: Employees
    def get_employees(self, status_filter=None):
        if status_filter == 'active':
            rows = db.query_db("SELECT * FROM employees WHERE status='active' ORDER BY name ASC")
        elif status_filter == 'inactive':
            rows = db.query_db("SELECT * FROM employees WHERE status='inactive' ORDER BY name ASC")
        else:
            rows = db.query_db("SELECT * FROM employees ORDER BY status ASC, name ASC")
        employees = [dict(r) for r in rows]
        self.send_json_response(employees)

    def add_employee(self, data):
        name = data.get('name', '').strip()
        phone = data.get('phone', '').strip()
        national_id = data.get('national_id', '').strip()
        pay_type = data.get('pay_type', 'hourly')
        rate = float(data.get('rate', 0))
        default_shift = data.get('default_shift', 'morning')
        shift_start_time = data.get('shift_start_time', '').strip()
        shift_end_time = data.get('shift_end_time', '').strip()
        
        if not name:
            self.send_json_response({"error": "الاسم مطلوب"}, 400)
            return

        if national_id:
            existing = db.query_db("SELECT id FROM employees WHERE national_id=? AND national_id!=''", (national_id,), one=True)
            if existing:
                self.send_json_response({"error": "رقم الهوية هذا مسجل بالفعل لموظف آخر"}, 400)
                return

        last = db.query_db("SELECT employee_code FROM employees WHERE employee_code LIKE 'EMP-%' ORDER BY id DESC LIMIT 1", one=True)
        if last and last['employee_code']:
            try:
                num = int(last['employee_code'].split('-')[1]) + 1
            except Exception:
                num = 1
        else:
            num = 1
        employee_code = 'EMP-' + str(num).zfill(3)

        db.execute_db('''
            INSERT INTO employees (name, employee_code, phone, national_id, pay_type, rate, default_shift, status, shift_start_time, shift_end_time)
            VALUES (?, ?, ?, ?, ?, ?, ?, 'active', ?, ?)
        ''', (name, employee_code, phone, national_id, pay_type, rate, default_shift, shift_start_time, shift_end_time))
        
        self.send_json_response({"success": True, "employee_code": employee_code})

    def update_employee(self, data):
        emp_id = data.get('id')
        name = data.get('name', '').strip()
        employee_code = data.get('employee_code', '').strip()
        phone = data.get('phone', '').strip()
        national_id = data.get('national_id', '').strip()
        pay_type = data.get('pay_type')
        rate = float(data.get('rate', 0))
        default_shift = data.get('default_shift')
        status = data.get('status', 'active')
        shift_start_time = data.get('shift_start_time', '').strip()
        shift_end_time = data.get('shift_end_time', '').strip()

        if not emp_id or not name:
            self.send_json_response({"error": "البيانات غير مكتملة"}, 400)
            return

        if national_id:
            existing = db.query_db("SELECT id FROM employees WHERE national_id=? AND id!=? AND national_id!=''", (national_id, emp_id), one=True)
            if existing:
                self.send_json_response({"error": "رقم الهوية هذا مسجل بالفعل لموظف آخر"}, 400)
                return

        db.execute_db('''
            UPDATE employees
            SET name=?, employee_code=?, phone=?, national_id=?, pay_type=?, rate=?, default_shift=?, status=?, shift_start_time=?, shift_end_time=?
            WHERE id=?
        ''', (name, employee_code, phone, national_id, pay_type, rate, default_shift, status, shift_start_time, shift_end_time, emp_id))
        
        self.send_json_response({"success": True})

    def delete_employee(self, data):
        emp_id = data.get('id')
        if not emp_id:
            self.send_json_response({"error": "رقم الموظف مطلوب"}, 400)
            return

        # Soft delete by setting status to inactive
        db.execute_db("UPDATE employees SET status='inactive' WHERE id=?", (emp_id,))
        self.send_json_response({"success": True})

    def restore_employee(self, data):
        emp_id = data.get('id')
        if not emp_id:
            self.send_json_response({"error": "رقم الموظف مطلوب"}, 400)
            return
        db.execute_db("UPDATE employees SET status='active' WHERE id=?", (emp_id,))
        self.send_json_response({"success": True})

    # API Handlers: Attendance
    def get_attendance(self, start_date, end_date):
        if not start_date or not end_date:
            end = datetime.date.today()
            start = end - datetime.timedelta(days=30)
            start_date = start.isoformat()
            end_date = end.isoformat()
            
        rows = db.query_db('''
            SELECT a.*, e.name as employee_name, e.pay_type, e.rate
            FROM attendance a 
            JOIN employees e ON a.employee_id = e.id
            WHERE a.date >= ? AND a.date <= ?
            ORDER BY a.date DESC, a.shift ASC, e.name ASC
        ''', (start_date, end_date))
        
        attendance = [dict(r) for r in rows]
        self.send_json_response(attendance)

    def log_attendance(self, data):
        employee_id = data.get('employee_id')
        date = data.get('date')
        shift = data.get('shift', 'morning')
        check_in = data.get('check_in', '08:00 AM')
        check_out = data.get('check_out', '04:00 PM')
        notes = data.get('notes', '')
        excuse = data.get('excuse', '')

        if not employee_id or not date:
            self.send_json_response({"error": "البيانات غير مكتملة"}, 400)
            return

        emp = db.query_db("SELECT * FROM employees WHERE id=?", (employee_id,), one=True)
        if not emp:
            self.send_json_response({"error": "الموظف غير موجود"}, 400)
            return

        shift_start, shift_end = get_employee_shift_times(emp)

        # Check for existing record for same employee/date/shift
        existing = db.query_db('''
            SELECT id, check_in, check_out FROM attendance 
            WHERE employee_id=? AND date=? AND shift=?
        ''', (employee_id, date, shift), one=True)

        if existing and existing['check_in'] and existing['check_out'] and existing['check_out'] != '00:00':
            self.send_json_response({"error": "تم تسجيل حضور وانصراف هذا الموظف بالفعل في هذا اليوم والوردية. يرجى حذف السجل الحالي أولاً قبل التسجيل مجدداً."}, 400)
            return

        try:
            in_mins = parse_12h_to_24h_minutes(check_in)
            out_mins = parse_12h_to_24h_minutes(check_out)
            diff_mins = out_mins - in_mins
            if diff_mins < 0:
                diff_mins += 1440
            hours_worked = round(diff_mins / 60.0, 2)
        except Exception:
            hours_worked = 8.0
            in_mins = shift_start
            out_mins = shift_end

        shift_duration_mins = shift_end - shift_start
        if shift_duration_mins <= 0:
            shift_duration_mins += 1440
        expected_hours = round(shift_duration_mins / 60.0, 2)
        overtime_hours = max(0.0, hours_worked - expected_hours)
        
        # Late calculation: compare check-in to configured shift start
        is_late = 0
        late_minutes = 0
        deduction_hours = 0
        if in_mins > shift_start:
            late_minutes = in_mins - shift_start
            is_late = 1
            if excuse:
                deduction_hours = round(late_minutes / 60.0, 2)
            else:
                deduction_hours = round(late_minutes / 30.0, 2)

        check_in_24 = minutes_to_hhmm(in_mins)
        check_out_24 = minutes_to_hhmm(out_mins)

        if existing:
            db.execute_db('''
                UPDATE attendance
                SET check_in=?, check_out=?, hours_worked=?, overtime_hours=?, notes=?, is_late=?, late_minutes=?, excuse=?, deduction_hours=?
                WHERE id=?
            ''', (check_in_24, check_out_24, hours_worked, overtime_hours, notes, is_late, late_minutes, excuse, deduction_hours, existing['id']))
        else:
            try:
                db.execute_db('''
                    INSERT INTO attendance (employee_id, date, shift, check_in, check_out, hours_worked, overtime_hours, notes, is_late, late_minutes, excuse, deduction_hours)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (employee_id, date, shift, check_in_24, check_out_24, hours_worked, overtime_hours, notes, is_late, late_minutes, excuse, deduction_hours))
            except Exception as e:
                self.send_json_response({"error": "يوجد سجل حضور مسجل بالفعل لهذا الموظف في هذا اليوم والوردية. يرجى حذف السجل الحالي أولاً."}, 400)
                return

        self.send_json_response({"success": True, "late_minutes": late_minutes, "deduction_hours": deduction_hours})

    def delete_attendance(self, data):
        att_id = data.get('id')
        if not att_id:
            self.send_json_response({"error": "رقم السجل مطلوب"}, 400)
            return
        db.execute_db("DELETE FROM attendance WHERE id=?", (att_id,))
        self.send_json_response({"success": True})

    def fast_check(self, data):
        employee_id = data.get('employee_id')
        now = datetime.datetime.now()
        date = data.get('date', now.strftime('%Y-%m-%d'))
        current_time = data.get('time', now.strftime('%H:%M'))
        notes = data.get('notes', '')

        if not employee_id:
            self.send_json_response({"error": "يجب اختيار الموظف"}, 400)
            return

        emp = db.query_db("SELECT * FROM employees WHERE id=?", (employee_id,), one=True)
        if not emp:
            self.send_json_response({"error": "الموظف غير موجود"}, 400)
            return

        shift = emp['default_shift']
        shift_start, shift_end = get_employee_shift_times(emp)

        existing = db.query_db('''
            SELECT id, check_in, check_out FROM attendance 
            WHERE employee_id=? AND date=? AND shift=?
        ''', (employee_id, date, shift), one=True)

        current_mins = parse_24h_to_minutes(current_time)

        if existing:
            if not existing['check_out'] or existing['check_out'] == '00:00':
                # Has check-in but no check-out yet -> record check-out
                in_mins = parse_24h_to_minutes(existing['check_in'])
                diff_mins = current_mins - in_mins
                if diff_mins < 0:
                    diff_mins += 1440
                hours_worked = round(diff_mins / 60.0, 2)
                shift_duration_mins = shift_end - shift_start
                if shift_duration_mins <= 0:
                    shift_duration_mins += 1440
                expected_hours = round(shift_duration_mins / 60.0, 2)
                overtime_hours = max(0.0, hours_worked - expected_hours)
                
                is_late = 0
                late_minutes = 0
                deduction_hours = 0
                if in_mins > shift_start:
                    late_minutes = in_mins - shift_start
                    is_late = 1
                
                db.execute_db('''
                    UPDATE attendance SET check_out=?, hours_worked=?, overtime_hours=?, notes=?, is_late=?, late_minutes=?, deduction_hours=?
                    WHERE id=?
                ''', (current_time, hours_worked, overtime_hours, notes or '', is_late, late_minutes, deduction_hours, existing['id']))
                self.send_json_response({"success": True, "action": "check_out", "time": minutes_to_12h(current_mins)})
            else:
                # Already has both check-in and check-out -> block duplicate
                self.send_json_response({"error": "تم تسجيل الحضور والانصراف لهذا اليوم بالفعل"}, 400)
        else:
            # No existing record -> create check-in with no check-out yet
            is_late = 0
            late_minutes = 0
            if current_mins > shift_start:
                late_minutes = current_mins - shift_start
                is_late = 1
            
            try:
                db.execute_db('''
                    INSERT INTO attendance (employee_id, date, shift, check_in, check_out, hours_worked, overtime_hours, notes, is_late, late_minutes, deduction_hours)
                    VALUES (?, ?, ?, ?, '00:00', 0.0, 0.0, ?, ?, ?, 0.0)
                ''', (employee_id, date, shift, current_time, notes or '', is_late, late_minutes))
            except Exception as e:
                self.send_json_response({"error": "يوجد سجل حضور مسجل بالفعل لهذا الموظف في هذا اليوم والوردية"}, 400)
                return
            self.send_json_response({"success": True, "action": "check_in", "time": minutes_to_12h(current_mins)})

    # API Handlers: Financial Transactions (Loans, Bonuses, Deductions)
    def get_transactions(self, employee_id):
        if employee_id:
            rows = db.query_db('''
                SELECT t.*, e.name as employee_name 
                FROM transactions t
                JOIN employees e ON t.employee_id = e.id
                WHERE t.employee_id = ?
                ORDER BY t.date DESC
            ''', (employee_id,))
        else:
            rows = db.query_db('''
                SELECT t.*, e.name as employee_name 
                FROM transactions t
                JOIN employees e ON t.employee_id = e.id
                ORDER BY t.date DESC
            ''')
            
        transactions = [dict(r) for r in rows]

        # Calculate active balance for each employee (optimized: 2 queries total)
        balances = {}
        loans_by_emp = {}
        for r in db.query_db("SELECT employee_id, SUM(amount) as sum FROM transactions WHERE type='loan' GROUP BY employee_id"):
            loans_by_emp[r['employee_id']] = r['sum'] or 0
        repaid_by_emp = {}
        for r in db.query_db("SELECT employee_id, SUM(amount) as sum FROM transactions WHERE type='deduction' AND description LIKE '%تسوية سلفة%' GROUP BY employee_id"):
            repaid_by_emp[r['employee_id']] = r['sum'] or 0
        for r in db.query_db("SELECT id, name FROM employees WHERE status='active'"):
            emp_id = r['id']
            balances[emp_id] = {
                "name": r['name'],
                "outstanding_loan": (loans_by_emp.get(emp_id, 0) or 0) - (repaid_by_emp.get(emp_id, 0) or 0)
            }

        self.send_json_response({
            "transactions": transactions,
            "balances": balances
        })

    def add_transaction(self, data):
        employee_id = data.get('employee_id')
        t_type = data.get('type')  # 'loan', 'bonus', 'deduction'
        amount = float(data.get('amount', 0))
        date = data.get('date')
        description = data.get('description', '')

        if not employee_id or not t_type or not amount or not date:
            self.send_json_response({"error": "البيانات غير مكتملة"}, 400)
            return

        emp = db.query_db("SELECT name FROM employees WHERE id=?", (employee_id,), one=True)
        emp_name = emp['name'] if emp else ''

        db.execute_db('''
            INSERT INTO transactions (employee_id, type, amount, date, description)
            VALUES (?, ?, ?, ?, ?)
        ''', (employee_id, t_type, amount, date, description))

        # If loan/advance: automatically deduct from treasury as expense
        if t_type == 'loan':
            desc = f'سلفة للموظف {emp_name}' + (f' - {description}' if description else '')
            self._treasury_add('withdrawal', amount, date, desc, 'مصروفات عامة', emp_name)

        self.send_json_response({"success": True})

    def delete_transaction(self, data):
        t_id = data.get('id')
        if not t_id:
            self.send_json_response({"error": "رقم الحركة مطلوب"}, 400)
            return

        tx = db.query_db("SELECT payroll_id, type, amount, employee_id, date, description FROM transactions WHERE id=?", (t_id,), one=True)

        if tx and tx['payroll_id']:
            self.send_json_response({"error": "لا يمكن حذف حركة مالية تم تسويتها في مسير رواتب مغلق"}, 400)
            return

        conn = db.get_conn()
        try:
            conn.execute("BEGIN IMMEDIATE")
            cursor = conn.cursor()

            if tx and tx['type'] == 'loan':
                rev_row = cursor.execute("SELECT id, type, amount FROM cashbox WHERE description LIKE ? ORDER BY id DESC LIMIT 1", ('%' + f'سلفة للموظف' + '%',)).fetchone()
                if rev_row:
                    rev_type = 'deposit' if rev_row['type'] == 'withdrawal' else 'withdrawal'
                    r = cursor.execute("SELECT COALESCE(SUM(CASE WHEN type='deposit' THEN amount ELSE 0 END), 0) as deposits, COALESCE(SUM(CASE WHEN type='withdrawal' THEN amount ELSE 0 END), 0) as withdrawals FROM cashbox", one=True).fetchone()
                    deposits, withdrawals = r['deposits'], r['withdrawals']
                    bal = deposits - withdrawals
                    if rev_type == 'deposit':
                        bal += rev_row['amount']
                    else:
                        bal -= rev_row['amount']
                    cursor.execute("INSERT INTO cashbox (date, type, amount, source, description, category, balance_after) VALUES (?, ?, ?, ?, ?, ?, ?)",
                                   (tx['date'], rev_type, rev_row['amount'], 'عكس حذف سلفة', 'عكس: سلفة للموظف', '', bal))

            cursor.execute("DELETE FROM transactions WHERE id=?", (t_id,))
            conn.commit()
            self.send_json_response({"success": True})
        except Exception as e:
            conn.rollback()
            server_log(f"delete_transaction failed, rolled back: {e}")
            self.send_json_response({"error": f"فشل حذف الحركة: {str(e)}"}, 500)

    # API Handlers: Inventory (BVC Stock)
    def get_inventory(self):
        search = None
        rows = db.query_db("SELECT * FROM inventory ORDER BY item_name ASC")
        items = [dict(r) for r in rows]
        self.send_json_response(items)

    def get_inventory_search(self, query):
        search = query.get('search', [None])[0]
        limit = query.get('limit', [None])[0]
        offset = query.get('offset', [None])[0]
        stock_filter = query.get('stock', [None])[0]

        base_sql = "SELECT * FROM inventory"
        count_sql = "SELECT COUNT(*) FROM inventory"
        params = []
        where_clauses = []

        if search:
            where_clauses.append("(item_name LIKE ? OR item_code LIKE ?)")
            params.extend(['%' + search + '%', '%' + search + '%'])

        if stock_filter == 'low':
            where_clauses.append("quantity <= min_stock")

        if where_clauses:
            where_sql = " WHERE " + " AND ".join(where_clauses)
            base_sql += where_sql
            count_sql += where_sql

        total = db.query_db(count_sql, tuple(params), one=True)[0]

        base_sql += " ORDER BY item_name ASC"
        if limit:
            base_sql += " LIMIT ?"
            params.append(int(limit))
        if offset:
            base_sql += " OFFSET ?"
            params.append(int(offset))

        rows = db.query_db(base_sql, tuple(params))
        items = [dict(r) for r in rows]
        self.send_json_response({"items": items, "total": total})

    def add_inventory_item(self, data):
        item_name = data.get('item_name', '').strip()
        item_code = data.get('item_code', '').strip()
        quantity = int(data.get('quantity', 0))
        unit = data.get('unit', 'قطعة').strip()
        purchase_price = float(data.get('purchase_price', 0))
        sale_price = float(data.get('sale_price', 0))
        min_stock = int(data.get('min_stock', 5))
        description = data.get('description', '').strip()
        barcode = data.get('barcode', '').strip()

        if not item_name or not item_code:
            self.send_json_response({"error": "الاسم والكود مطلوبان"}, 400)
            return

        try:
            db.execute_db('''
                INSERT INTO inventory (item_name, item_code, quantity, unit, purchase_price, sale_price, min_stock, description, barcode)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (item_name, item_code, quantity, unit, purchase_price, sale_price, min_stock, description, barcode))
            self.send_json_response({"success": True})
        except sqlite3.IntegrityError:
            self.send_json_response({"error": "كود الصنف مسجل مسبقاً لمادة أخرى"}, 400)

    def update_inventory_item(self, data):
        item_id = data.get('id')
        item_name = data.get('item_name', '').strip()
        item_code = data.get('item_code', '').strip()
        quantity = int(data.get('quantity', 0))
        unit = data.get('unit', 'قطعة').strip()
        purchase_price = float(data.get('purchase_price', 0))
        sale_price = float(data.get('sale_price', 0))
        min_stock = int(data.get('min_stock', 5))
        description = data.get('description', '').strip()
        barcode = data.get('barcode', '').strip()

        if not item_id or not item_name or not item_code:
            self.send_json_response({"error": "البيانات غير مكتملة"}, 400)
            return

        try:
            db.execute_db('''
                UPDATE inventory
                SET item_name=?, item_code=?, quantity=?, unit=?, purchase_price=?, sale_price=?, min_stock=?, description=?, barcode=?, updated_at=CURRENT_TIMESTAMP
                WHERE id=?
            ''', (item_name, item_code, quantity, unit, purchase_price, sale_price, min_stock, description, barcode, item_id))
            self.send_json_response({"success": True})
        except sqlite3.IntegrityError:
            self.send_json_response({"error": "كود الصنف مسجل مسبقاً لمادة أخرى"}, 400)

    def delete_inventory_item(self, data):
        item_id = data.get('id')
        if not item_id:
            self.send_json_response({"error": "رقم الصنف مطلوب"}, 400)
            return
        db.execute_db("DELETE FROM inventory WHERE id=?", (item_id,))
        self.send_json_response({"success": True})

    # Global Search
    def global_search(self, query):
        q = query.get('q', [''])[0].strip()
        if not q or len(q) < 1:
            self.send_json_response({"employees": [], "inventory": [], "transactions": []})
            return
        like = '%' + q + '%'
        employees = [dict(r) for r in db.query_db(
            "SELECT id, name, employee_code, phone, national_id FROM employees WHERE name LIKE ? OR employee_code LIKE ? OR phone LIKE ? OR national_id LIKE ? LIMIT 20",
            (like, like, like, like))]
        inventory = [dict(r) for r in db.query_db(
            "SELECT id, item_name, item_code, barcode, quantity, sale_price FROM inventory WHERE item_name LIKE ? OR item_code LIKE ? OR barcode LIKE ? LIMIT 20",
            (like, like, like))]
        transactions = [dict(r) for r in db.query_db(
            "SELECT t.id, t.type, t.amount, t.date, t.description, e.name as employee_name FROM transactions t JOIN employees e ON t.employee_id = e.id WHERE e.name LIKE ? OR t.description LIKE ? OR CAST(t.amount AS TEXT) LIKE ? LIMIT 20",
            (like, like, like))]
        attendance = [dict(r) for r in db.query_db(
            "SELECT a.id, a.date, a.check_in, a.check_out, e.name as employee_name FROM attendance a JOIN employees e ON a.employee_id = e.id WHERE e.name LIKE ? OR a.date LIKE ? LIMIT 20",
            (like, like))]
        self.send_json_response({"employees": employees, "inventory": inventory, "transactions": transactions, "attendance": attendance})

    # Barcode Lookup
    def barcode_lookup(self, query):
        barcode = query.get('barcode', [''])[0].strip()
        if not barcode:
            self.send_json_response({"error": "الباركود مطلوب"}, 400)
            return
        item = db.query_db("SELECT * FROM inventory WHERE barcode=? LIMIT 1", (barcode,), one=True)
        if item:
            self.send_json_response(dict(item))
        else:
            self.send_json_response({"error": "الصنف غير موجود"}, 404)

    def barcode_update(self, data):
        item_id = data.get('id')
        barcode = data.get('barcode', '').strip()
        if not item_id:
            self.send_json_response({"error": "رقم الصنف مطلوب"}, 400)
            return
        db.execute_db("UPDATE inventory SET barcode=? WHERE id=?", (barcode, item_id))
        self.send_json_response({"success": True})

    # Printable Receipts
    def receipt_attendance(self, query):
        attendance_id = query.get('id', [None])[0]
        if not attendance_id:
            self.send_json_response({"error": "رقم السجل مطلوب"}, 400)
            return
        row = db.query_db('''
            SELECT a.*, e.name as employee_name, e.employee_code
            FROM attendance a JOIN employees e ON a.employee_id = e.id
            WHERE a.id=?
        ''', (attendance_id,), one=True)
        if not row:
            self.send_json_response({"error": "السجل غير موجود"}, 404)
            return
        settings = self._get_settings_dict()
        company = settings.get('company_name', 'شركة الرحمه')
        html = self._render_receipt_html("إذن حضور وانصراف", company, [
            ("الموظف", row['employee_name']),
            ("كود الموظف", row['employee_code']),
            ("التاريخ", row['date']),
            ("الوردية", row['shift']),
            ("حضور", row['check_in']),
            ("انصراف", row['check_out'] if row['check_out'] != '00:00' else '---'),
            ("ساعات العمل", str(row['hours_worked'])),
            ("ساعات إضافي", str(row['overtime_hours'])),
        ])
        self.send_response(200)
        self.send_header('Content-Type', 'text/html; charset=utf-8')
        self.send_header('Content-Length', str(len(html.encode('utf-8'))))
        self.end_headers()
        self.wfile.write(html.encode('utf-8'))

    def receipt_payroll(self, query):
        payroll_id = query.get('id', [None])[0]
        if not payroll_id:
            self.send_json_response({"error": "رقم المسير مطلوب"}, 400)
            return
        row = db.query_db('''
            SELECT p.*, e.name as employee_name, e.employee_code
            FROM payroll p JOIN employees e ON p.employee_id = e.id
            WHERE p.id=?
        ''', (payroll_id,), one=True)
        if not row:
            self.send_json_response({"error": "المسير غير موجود"}, 404)
            return
        settings = self._get_settings_dict()
        company = settings.get('company_name', 'شركة الرحمه')
        html = self._render_receipt_html("سند صرف راتب", company, [
            ("الموظف", row['employee_name']),
            ("كود الموظف", row['employee_code']),
            ("من تاريخ", row['start_date']),
            ("إلى تاريخ", row['end_date']),
            ("الراتب الأساسي", str(row['base_salary'])),
            ("المكافآت", str(row['total_bonuses'])),
            ("الخصومات", str(row['total_deductions'])),
            ("خصم السلف", str(row['total_loans_deducted'])),
            ("صافي الراتب", str(row['net_salary'])),
        ])
        self.send_response(200)
        self.send_header('Content-Type', 'text/html; charset=utf-8')
        self.send_header('Content-Length', str(len(html.encode('utf-8'))))
        self.end_headers()
        self.wfile.write(html.encode('utf-8'))

    def _get_settings_dict(self):
        rows = db.query_db("SELECT key, value FROM system_settings")
        return {r['key']: r['value'] for r in rows}

    def _render_receipt_html(self, title, company, fields):
        rows_html = ''
        for label, val in fields:
            rows_html += f'<tr><td style="padding:8px 12px;border:1px solid #ccc;font-weight:700;background:#f8fafc">{label}</td><td style="padding:8px 12px;border:1px solid #ccc;text-align:left">{val}</td></tr>\n'
        return f'''<!DOCTYPE html><html dir="rtl"><head><meta charset="utf-8"><title>{title}</title>
<style>body{{font-family:Tajawal,'Segoe UI',sans-serif;padding:40px;max-width:400px;margin:auto}}
h2{{text-align:center;color:#0d9488;margin-bottom:4px}}
.sub{{text-align:center;color:#64748b;font-size:0.85rem;margin-bottom:20px}}
table{{width:100%;border-collapse:collapse}}
@media print{{body{{padding:20px}}.no-print{{display:none}}}}
</style></head><body>
<h2>{company}</h2>
<p class="sub">{title}</p>
<table>{rows_html}</table>
<div style="text-align:center;margin-top:24px" class="no-print">
<button onclick="window.print()" style="padding:10px 24px;background:#0d9488;color:#fff;border:none;border-radius:8px;cursor:pointer;font-size:1rem;font-weight:700">طباعة</button>
</div></body></html>'''

    # API Handlers: Payroll Calculation
    def calculate_payroll(self, start_date, end_date):
        if not start_date or not end_date:
            self.send_json_response({"error": "تاريخ البدء والنهاية مطلوبان"}, 400)
            return

        # Fetch active employees
        employees = db.query_db("SELECT * FROM employees WHERE status='active'")
        payroll_records = []

        for emp in employees:
            emp_id = emp['id']
            pay_type = emp['pay_type']
            rate = emp['rate']

            # Get attendance for this employee in the date range
            attendance = db.query_db('''
                SELECT * FROM attendance 
                WHERE employee_id=? AND date>=? AND date<=?
            ''', (emp_id, start_date, end_date))

            # Compute hours, shifts, and base salary
            total_hours = 0.0
            total_shifts = len(attendance)
            overtime_hours = 0.0
            overtime_pay = 0.0

            for att in attendance:
                total_hours += att['hours_worked']
                overtime_hours += att['overtime_hours']

            if pay_type == 'hourly':
                base_salary = round(total_hours * rate, 2)
            else:  # 'shift'
                # Fixed rate per shift, plus overtime pay for extra hours (base rate/8 per hour)
                base_salary = round(total_shifts * rate, 2)
                # Overtime pay: (rate / 8) * overtime_hours
                overtime_pay = round((rate / 8.0) * overtime_hours, 2)
                base_salary += overtime_pay

            # Get unsettled transactions within the payroll period for this employee
            unsettled_tx = db.query_db('''
                SELECT * FROM transactions 
                WHERE employee_id=? AND payroll_id IS NULL AND date >= ? AND date <= ?
            ''', (emp_id, start_date, end_date))

            total_bonuses = 0.0
            total_deductions = 0.0
            
            # Loans taken
            loans_taken_in_period = 0.0
            for tx in unsettled_tx:
                if tx['type'] == 'bonus':
                    total_bonuses += tx['amount']
                elif tx['type'] == 'deduction':
                    total_deductions += tx['amount']
                elif tx['type'] == 'loan':
                    # A loan taken is outstanding. It is not directly deducted,
                    # but we track it so the user can choose how much to deduct.
                    loans_taken_in_period += tx['amount']

            # Get overall outstanding loan balance
            total_loans_taken = db.query_db("SELECT SUM(amount) as sum FROM transactions WHERE employee_id=? AND type='loan'", (emp_id,), one=True)['sum'] or 0
            total_loans_repaid = db.query_db("SELECT SUM(amount) as sum FROM transactions WHERE employee_id=? AND type='deduction' AND description LIKE '%تسوية سلفة%'", (emp_id,), one=True)['sum'] or 0
            outstanding_loan_balance = total_loans_taken - total_loans_repaid

            payroll_records.append({
                "employee_id": emp_id,
                "name": emp['name'],
                "pay_type": pay_type,
                "rate": rate,
                "total_hours": total_hours,
                "total_shifts": total_shifts,
                "overtime_hours": overtime_hours,
                "overtime_pay": overtime_pay,
                "base_salary": base_salary,
                "total_bonuses": total_bonuses,
                "total_deductions": total_deductions,
                "outstanding_loan_balance": outstanding_loan_balance,
                "suggested_loan_deduction": min(outstanding_loan_balance, base_salary),
                "net_salary": base_salary + total_bonuses - total_deductions
            })

        self.send_json_response({
            "start_date": start_date,
            "end_date": end_date,
            "records": payroll_records
        })

    def save_payroll(self, data):
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        records = data.get('records', [])

        if not start_date or not end_date or not records:
            self.send_json_response({"error": "البيانات غير كاملة"}, 400)
            return

        conn = db.get_conn()
        try:
            conn.execute("BEGIN IMMEDIATE")
            cursor = conn.cursor()
            for rec in records:
                emp_id = rec.get('employee_id')
                total_hours = float(rec.get('total_hours', 0))
                total_shifts = int(rec.get('total_shifts', 0))
                base_salary = float(rec.get('base_salary', 0))
                total_bonuses = float(rec.get('total_bonuses', 0))
                total_deductions = float(rec.get('total_deductions', 0))
                loan_deduction = float(rec.get('loan_deduction', 0))
                net_salary = float(rec.get('net_salary', 0))

                emp_row = cursor.execute("SELECT name FROM employees WHERE id=?", (emp_id,)).fetchone()
                emp_name = emp_row['name'] if emp_row else ''

                cursor.execute('''
                    INSERT INTO payroll (employee_id, start_date, end_date, total_hours, total_shifts, base_salary, total_bonuses, total_deductions, total_loans_deducted, net_salary)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (emp_id, start_date, end_date, total_hours, total_shifts, base_salary, total_bonuses, total_deductions, loan_deduction, net_salary))
                payroll_id = cursor.lastrowid

                cursor.execute('''
                    UPDATE transactions SET payroll_id = ? WHERE employee_id = ? AND payroll_id IS NULL AND date >= ? AND date <= ?
                ''', (payroll_id, emp_id, start_date, end_date))

                if loan_deduction > 0:
                    cursor.execute('''
                        INSERT INTO transactions (employee_id, type, amount, date, description, payroll_id)
                        VALUES (?, 'deduction', ?, ?, 'تسوية سلفة - كشف رواتب أسبوعي', ?)
                    ''', (emp_id, loan_deduction, end_date, payroll_id))
                    deposits, withdrawals, balance = self._get_treasury_totals()
                    balance += loan_deduction
                    cursor.execute('''
                        INSERT INTO cashbox (date, type, amount, source, description, category, balance_after)
                        VALUES (?, 'deposit', ?, ?, ?, 'رواتب', ?)
                    ''', (end_date, loan_deduction, f'تسوية سلفة من {emp_name} - كشف رواتب', f'تسوية سلفة من {emp_name} - كشف رواتب', balance))

                if net_salary > 0:
                    deposits, withdrawals, balance = self._get_treasury_totals()
                    balance -= net_salary
                    cursor.execute('''
                        INSERT INTO cashbox (date, type, amount, source, description, category, balance_after)
                        VALUES (?, 'withdrawal', ?, ?, ?, 'رواتب', ?)
                    ''', (end_date, net_salary, f'رواتب {emp_name} ({start_date} - {end_date})', f'رواتب {emp_name} ({start_date} - {end_date})', balance))

            conn.commit()
            self.send_json_response({"success": True})
        except Exception as e:
            conn.rollback()
            server_log(f"save_payroll failed, rolled back: {e}")
            self.send_json_response({"error": f"فشل حفظ كشف الرواتب: {str(e)}"}, 500)

    # Export CSV for Payroll
    def export_payroll_csv(self, start_date, end_date):
        if not start_date or not end_date:
            self.send_response(400)
            self.end_headers()
            self.wfile.write(b"Error: Start and end dates are required")
            return

        # Fetch saved payroll records in this date range
        rows = db.query_db('''
            SELECT p.*, e.name as employee_name, e.pay_type, e.rate
            FROM payroll p
            JOIN employees e ON p.employee_id = e.id
            WHERE p.start_date = ? AND p.end_date = ?
            ORDER BY e.name ASC
        ''', (start_date, end_date))

        # Calculate live and build the CSV
        employees = db.query_db("SELECT * FROM employees WHERE status='active'")
        csv_data = []
        
        # CSV Header
        csv_data.append([
            "اسم الموظف",
            "نوع الحساب",
            "سعر الساعة/الوردية",
            "الساعات الفعلية",
            "عدد الورديات",
            "الراتب الأساسي (شامل الإضافي)",
            "إجمالي المكافآت",
            "إجمالي الخصومات",
            "المستقطع للسلف",
            "صافي الراتب المستحق",
            "توقيع المستلم"
        ])

        for emp in employees:
            emp_id = emp['id']
            pay_type = emp['pay_type']
            rate = emp['rate']

            attendance = db.query_db('''
                SELECT * FROM attendance 
                WHERE employee_id=? AND date>=? AND date<=?
            ''', (emp_id, start_date, end_date))

            total_hours = 0.0
            total_shifts = len(attendance)
            overtime_hours = 0.0
            overtime_pay = 0.0

            for att in attendance:
                total_hours += att['hours_worked']
                overtime_hours += att['overtime_hours']

            if pay_type == 'hourly':
                base_salary = round(total_hours * rate, 2)
                pay_type_ar = "بالساعة"
            else:
                base_salary = round(total_shifts * rate, 2)
                overtime_pay = round((rate / 8.0) * overtime_hours, 2)
                base_salary += overtime_pay
                pay_type_ar = "بالوردية"

            unsettled_tx = db.query_db('''
                SELECT * FROM transactions 
                WHERE employee_id=? AND payroll_id IS NULL AND date <= ?
            ''', (emp_id, end_date))

            total_bonuses = 0.0
            total_deductions = 0.0
            for tx in unsettled_tx:
                if tx['type'] == 'bonus':
                    total_bonuses += tx['amount']
                elif tx['type'] == 'deduction':
                    total_deductions += tx['amount']

            total_loans_taken = db.query_db("SELECT SUM(amount) as sum FROM transactions WHERE employee_id=? AND type='loan'", (emp_id,), one=True)['sum'] or 0
            total_loans_repaid = db.query_db("SELECT SUM(amount) as sum FROM transactions WHERE employee_id=? AND type='deduction' AND description LIKE '%تسوية سلفة%'", (emp_id,), one=True)['sum'] or 0
            outstanding_loan_balance = total_loans_taken - total_loans_repaid

            # Check if there is a saved record to get the exact loan deduction that was locked in
            saved = db.query_db('''
                SELECT * FROM payroll 
                WHERE employee_id=? AND start_date=? AND end_date=?
            ''', (emp_id, start_date, end_date), one=True)

            if saved:
                loan_deduction = saved['total_loans_deducted']
                net_salary = saved['net_salary']
                base_salary = saved['base_salary']
                total_bonuses = saved['total_bonuses']
                total_deductions = saved['total_deductions']
            else:
                # Suggest automatically deducting up to the net salary
                loan_deduction = min(outstanding_loan_balance, base_salary + total_bonuses - total_deductions)
                net_salary = base_salary + total_bonuses - total_deductions - loan_deduction

            csv_data.append([
                emp['name'],
                pay_type_ar,
                f"{rate} ج.م",
                f"{total_hours:.1f}" if pay_type == 'hourly' else "-",
                f"{total_shifts}" if pay_type == 'shift' else "-",
                f"{base_salary:.2f} ج.م",
                f"{total_bonuses:.2f} ج.م",
                f"{total_deductions:.2f} ج.م",
                f"{loan_deduction:.2f} ج.م",
                f"{net_salary:.2f} ج.م",
                "........................"
            ])

        # Generate CSV string
        import io
        output_stream = io.StringIO()
        writer = csv.writer(output_stream)
        writer.writerows(csv_data)
        
        # Convert to UTF-8 with BOM (utf-8-sig) for Arabic support in Excel
        csv_bytes = output_stream.getvalue().encode('utf-8-sig')

        filename = f"Payroll_BVC_{start_date}_to_{end_date}.csv"
        
        self.send_response(200)
        self.send_header('Content-Type', 'text/csv; charset=utf-8-sig')
        self.send_header('Content-Disposition', f'attachment; filename="{urllib.parse.quote(filename)}"')
        self.send_header('Content-Length', str(len(csv_bytes)))
        self.end_headers()
        self.wfile.write(csv_bytes)

    # API Handlers: Treasury
    def get_treasury(self, type_filter=None, date_from=None, date_to=None):
        query = "SELECT * FROM cashbox WHERE 1=1"
        params = []
        if type_filter:
            query += " AND type=?"
            params.append(type_filter)
        if date_from:
            query += " AND date>=?"
            params.append(date_from)
        if date_to:
            query += " AND date<=?"
            params.append(date_to)
        query += " ORDER BY date DESC, id DESC"
        rows = db.query_db(query, tuple(params))
        self.send_json_response([dict(r) for r in rows])

    def _get_treasury_totals(self):
        r = db.query_db("SELECT COALESCE(SUM(CASE WHEN type='deposit' THEN amount ELSE 0 END), 0) as deposits, COALESCE(SUM(CASE WHEN type='withdrawal' THEN amount ELSE 0 END), 0) as withdrawals FROM cashbox", one=True)
        return r['deposits'], r['withdrawals'], r['deposits'] - r['withdrawals']

    def get_treasury_balance(self):
        deposits, withdrawals, balance = self._get_treasury_totals()
        self.send_json_response({"balance": balance, "deposits": deposits, "withdrawals": withdrawals})

    def add_treasury_deposit(self, data):
        date = data.get('date', datetime.date.today().isoformat())
        amount = float(data.get('amount', 0))
        source = data.get('source', '')
        description = data.get('description', '')
        if amount <= 0:
            self.send_json_response({"error": "المبلغ يجب أن يكون أكبر من صفر"}, 400)
            return
        deposits, withdrawals, balance = self._get_treasury_totals()
        balance += amount
        db.execute_db('''
            INSERT INTO cashbox (date, type, amount, source, description, category, balance_after)
            VALUES (?, 'deposit', ?, ?, ?, 'وارد', ?)
        ''', (date, amount, source, description, balance))
        self.send_json_response({"success": True, "balance": balance})

    def add_treasury_withdrawal(self, data):
        date = data.get('date', datetime.date.today().isoformat())
        amount = float(data.get('amount', 0))
        description = data.get('description', '')
        category = data.get('category', 'مصروفات عامة')
        if amount <= 0:
            self.send_json_response({"error": "المبلغ يجب أن يكون أكبر من صفر"}, 400)
            return
        deposits, withdrawals, balance = self._get_treasury_totals()
        balance -= amount
        if balance < 0:
            self.send_json_response({"error": "الرصيد غير كافٍ"}, 400)
            return
        cashbox_id = db.execute_db('''
            INSERT INTO cashbox (date, type, amount, source, description, category, balance_after)
            VALUES (?, 'withdrawal', ?, ?, ?, ?, ?)
        ''', (date, amount, description, description, category, balance))
        db.execute_db('''
            INSERT INTO factory_expenses (date, amount, category, description, cashbox_id)
            VALUES (?, ?, ?, ?, ?)
        ''', (date, amount, category, description, cashbox_id))
        self.send_json_response({"success": True, "balance": balance})

    def _treasury_add(self, tx_type, amount, date, description, category='', source=''):
        deposits, withdrawals, balance = self._get_treasury_totals()
        if tx_type == 'deposit':
            balance += amount
        else:
            balance -= amount
        db.execute_db('''
            INSERT INTO cashbox (date, type, amount, source, description, category, balance_after)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (date, tx_type, amount, source or description, description, category, balance))
        return balance

    def _treasury_reverse(self, description_contains, date):
        row = db.query_db("SELECT id, type, amount FROM cashbox WHERE description LIKE ? ORDER BY id DESC LIMIT 1", ('%' + description_contains + '%',), one=True)
        if not row:
            return False
        reverse_type = 'deposit' if row['type'] == 'withdrawal' else 'withdrawal'
        self._treasury_add(reverse_type, row['amount'], date, 'عكس: ' + description_contains)
        return True

    # API Handlers: Employee Ledger
    def get_employee_ledger(self, employee_id):
        if employee_id:
            rows = db.query_db('''
                SELECT l.*, e.name as employee_name 
                FROM employee_ledger l
                JOIN employees e ON l.employee_id = e.id
                WHERE l.employee_id = ?
                ORDER BY l.date DESC, l.id DESC
            ''', (employee_id,))
        else:
            rows = db.query_db('''
                SELECT l.*, e.name as employee_name 
                FROM employee_ledger l
                JOIN employees e ON l.employee_id = e.id
                ORDER BY l.date DESC, l.id DESC
            ''')
        self.send_json_response([dict(r) for r in rows])

    def add_employee_ledger_entry(self, data):
        employee_id = data.get('employee_id')
        date = data.get('date', datetime.date.today().isoformat())
        entry_type = data.get('type')  # 'withdrawal', 'deduction', 'bonus', 'monthly_incentive', 'debt_add', 'debt_repayment'
        amount = float(data.get('amount', 0))
        description = data.get('description', '')
        if not employee_id or not entry_type or amount == 0:
            self.send_json_response({"error": "البيانات غير مكتملة"}, 400)
            return

        emp = db.query_db("SELECT name FROM employees WHERE id=?", (employee_id,), one=True)
        emp_name = emp['name'] if emp else ''

        db.execute_db('''
            INSERT INTO employee_ledger (employee_id, date, type, amount, description)
            VALUES (?, ?, ?, ?, ?)
        ''', (employee_id, date, entry_type, amount, description))

        # Auto-sync with treasury for direct cash movements
        if entry_type in ('withdrawal', 'debt_add'):
            # Money leaving treasury to employee
            self._treasury_add('withdrawal', abs(amount), date, f'{emp_name} - {description or entry_type}', 'رواتب', emp_name)
        elif entry_type == 'debt_repayment':
            # Money returning to treasury from employee
            self._treasury_add('deposit', abs(amount), date, f'تسوية دين من {emp_name} - {description or ""}', 'رواتب', emp_name)

        self.send_json_response({"success": True})

    # API Handlers: Factory Expenses
    def get_factory_expenses(self):
        rows = db.query_db("SELECT * FROM factory_expenses ORDER BY date DESC, id DESC")
        self.send_json_response([dict(r) for r in rows])

    def add_factory_expense(self, data):
        date = data.get('date', datetime.date.today().isoformat())
        amount = float(data.get('amount', 0))
        category = data.get('category', '')
        description = data.get('description', '')
        if amount <= 0:
            self.send_json_response({"error": "المبلغ يجب أن يكون أكبر من صفر"}, 400)
            return
        db.execute_db('''
            INSERT INTO factory_expenses (date, amount, category, description)
            VALUES (?, ?, ?, ?)
        ''', (date, amount, category, description))
        self.send_json_response({"success": True})

    # ==================== AUTH & USER MANAGEMENT ====================
    def _hash_password(self, password, salt=None):
        if not salt:
            salt = secrets.token_hex(16)
        pw_hash = hashlib.sha256((salt + password).encode('utf-8')).hexdigest()
        return salt + ':' + pw_hash

    def _verify_password(self, password, stored_hash):
        try:
            salt, pw_hash = stored_hash.split(':', 1)
            return hashlib.sha256((salt + password).encode('utf-8')).hexdigest() == pw_hash
        except Exception:
            return False

    def login_user(self, data):
        username = data.get('username', '').strip()
        password = data.get('password', '')
        if not username or not password:
            self.send_json_response({"error": "اسم المستخدم وكلمة المرور مطلوبان"}, 400)
            return
        user = db.query_db("SELECT * FROM system_users WHERE username=? AND is_active=1", (username,), one=True)
        if not user or not self._verify_password(password, user['password_hash']):
            self.send_json_response({"error": "بيانات الدخول غير صحيحة"}, 401)
            return
        db.execute_db("UPDATE system_users SET last_login=CURRENT_TIMESTAMP WHERE id=?", (user['id'],))
        user_dict = dict(user)
        perms = self._parse_permissions(user_dict.get('permissions'))
        self.send_json_response({"success": True, "user": {"id": user['id'], "username": user['username'], "display_name": user['display_name'], "role": user['role'], "permissions": perms}})

    def get_users(self):
        rows = db.query_db("SELECT id, username, display_name, role, is_active, permissions, created_at, last_login FROM system_users ORDER BY id ASC")
        result = []
        for r in rows:
            u = dict(r)
            u['permissions'] = self._parse_permissions(u.get('permissions'))
            result.append(u)
        self.send_json_response(result)

    def add_user(self, data):
        username = data.get('username', '').strip()
        password = data.get('password', '')
        display_name = data.get('display_name', '').strip()
        role = data.get('role', 'user')
        if not username or not password:
            self.send_json_response({"error": "اسم المستخدم وكلمة المرور مطلوبان"}, 400)
            return
        existing = db.query_db("SELECT id FROM system_users WHERE username=?", (username,), one=True)
        if existing:
            self.send_json_response({"error": "اسم المستخدم مسجل بالفعل"}, 400)
            return
        pw_hash = self._hash_password(password)
        db.execute_db(
            "INSERT INTO system_users (username, password_hash, display_name, role) VALUES (?, ?, ?, ?)",
            (username, pw_hash, display_name, role)
        )
        self.send_json_response({"success": True})

    def update_user(self, data):
        user_id = data.get('id')
        if not user_id:
            self.send_json_response({"error": "رقم المستخدم مطلوب"}, 400)
            return
        display_name = data.get('display_name', '').strip()
        role = data.get('role', 'user')
        is_active = data.get('is_active', 1)
        new_username = data.get('username', '').strip()
        if new_username:
            current = db.query_db("SELECT username FROM system_users WHERE id=?", (user_id,), one=True)
            if current:
                if current['username'] != new_username:
                    existing = db.query_db("SELECT id FROM system_users WHERE username=? AND id!=?", (new_username, user_id), one=True)
                    if existing:
                        self.send_json_response({"error": "اسم المستخدم مسجل بالفعل"}, 400)
                        return
                    db.execute_db("UPDATE system_users SET username=?, display_name=?, role=?, is_active=? WHERE id=?", (new_username, display_name, role, is_active, user_id))
                else:
                    db.execute_db("UPDATE system_users SET display_name=?, role=?, is_active=? WHERE id=?", (display_name, role, is_active, user_id))
            else:
                db.execute_db("UPDATE system_users SET display_name=?, role=?, is_active=? WHERE id=?", (display_name, role, is_active, user_id))
        else:
            db.execute_db("UPDATE system_users SET display_name=?, role=?, is_active=? WHERE id=?", (display_name, role, is_active, user_id))
        self.send_json_response({"success": True})

    def delete_user(self, data):
        user_id = data.get('id')
        if not user_id:
            self.send_json_response({"error": "رقم المستخدم مطلوب"}, 400)
            return
        user = db.query_db("SELECT username FROM system_users WHERE id=?", (user_id,), one=True)
        if user and user['username'] == 'admin':
            self.send_json_response({"error": "لا يمكن حذف حساب المدير الرئيسي"}, 400)
            return
        db.execute_db("DELETE FROM system_users WHERE id=?", (user_id,))
        self.send_json_response({"success": True})

    def reset_user_password(self, data):
        user_id = data.get('id')
        new_password = data.get('password', '')
        if not user_id or not new_password:
            self.send_json_response({"error": "رقم المستخدم وكلمة المرور الجديدة مطلوبان"}, 400)
            return
        pw_hash = self._hash_password(new_password)
        db.execute_db("UPDATE system_users SET password_hash=? WHERE id=?", (pw_hash, user_id))
        self.send_json_response({"success": True})

    def _parse_permissions(self, perms_json):
        """Parse the permissions JSON string, return dict with defaults for missing keys."""
        all_tabs = ['dashboard', 'employees', 'attendance', 'finance', 'payroll', 'inventory', 'treasury', 'settings']
        default = {t: True for t in all_tabs}
        if not perms_json:
            return default
        try:
            stored = json.loads(perms_json)
            for t in all_tabs:
                if t not in stored:
                    stored[t] = True
            return stored
        except (json.JSONDecodeError, TypeError):
            return default

    def get_user_permissions(self, query):
        user_id = query.get('user_id', [None])[0]
        if not user_id:
            self.send_json_response({"error": "رقم المستخدم مطلوب"}, 400)
            return
        user = db.query_db("SELECT id, username, display_name, role, permissions FROM system_users WHERE id=?", (user_id,), one=True)
        if not user:
            self.send_json_response({"error": "المستخدم غير موجود"}, 404)
            return
        perms = self._parse_permissions(user['permissions'])
        # Admin always gets all permissions
        if user['role'] == 'admin':
            perms = {t: True for t in perms}
        self.send_json_response({"user_id": user['id'], "username": user['username'], "permissions": perms})

    def save_user_permissions(self, data):
        user_id = data.get('user_id')
        permissions = data.get('permissions')
        if not user_id or not permissions:
            self.send_json_response({"error": "بيانات غير مكتملة"}, 400)
            return
        user = db.query_db("SELECT role FROM system_users WHERE id=?", (user_id,), one=True)
        if not user:
            self.send_json_response({"error": "المستخدم غير موجود"}, 404)
            return
        all_tabs = ['dashboard', 'employees', 'attendance', 'finance', 'payroll', 'inventory', 'treasury', 'settings']
        clean = {t: bool(permissions.get(t, True)) for t in all_tabs}
        db.execute_db("UPDATE system_users SET permissions=? WHERE id=?", (json.dumps(clean, ensure_ascii=False), user_id))
        self.send_json_response({"success": True})

    # ==================== SYSTEM SETTINGS ====================
    def get_settings(self):
        rows = db.query_db("SELECT key, value FROM system_settings")
        settings = {r['key']: r['value'] for r in rows}
        self.send_json_response(settings)

    def save_settings(self, data):
        for key, value in data.items():
            db.execute_db(
                "INSERT OR REPLACE INTO system_settings (key, value, updated_at) VALUES (?, ?, CURRENT_TIMESTAMP)",
                (key, str(value))
            )
        self.send_json_response({"success": True})

    def get_shifts(self):
        rows = db.query_db("SELECT id, name, start_time, end_time FROM shifts ORDER BY id ASC")
        self.send_json_response([dict(r) for r in rows])

    def save_shifts(self, data):
        shifts = data if isinstance(data, list) else data.get('shifts', [])
        for shift in shifts:
            sid = shift.get('id')
            name = shift.get('name', '').strip()
            start_time = shift.get('start_time', '').strip()
            end_time = shift.get('end_time', '').strip()
            if not name or not start_time or not end_time:
                continue
            if sid:
                db.execute_db("UPDATE shifts SET name=?, start_time=?, end_time=? WHERE id=?", (name, start_time, end_time, int(sid)))
            else:
                db.execute_db("INSERT INTO shifts (name, start_time, end_time) VALUES (?, ?, ?)", (name, start_time, end_time))
        self.send_json_response({"success": True})

    def delete_shift(self, data):
        sid = data.get('id')
        if not sid:
            self.send_json_response({"error": "رقم الوردية مطلوب"}, 400)
            return
        emp_refs = db.query_db("SELECT COUNT(*) as cnt FROM employees WHERE default_shift=?", (str(sid),), one=True)
        if emp_refs and emp_refs['cnt'] > 0:
            self.send_json_response({"error": "لا يمكن حذف الوردية لأنها مستخدمة من قبل الموظفين"}, 400)
            return
        db.execute_db("DELETE FROM shifts WHERE id=?", (int(sid),))
        self.send_json_response({"success": True})

    # ==================== BACKUP / RESTORE ====================
    def _backup_dir(self):
        d = os.path.join(BASE_DIR, 'backups')
        if not os.path.exists(d):
            os.makedirs(d)
        return d

    def list_backups(self):
        d = self._backup_dir()
        backups = []
        for f in sorted(os.listdir(d), reverse=True):
            if f.endswith('.db'):
                path = os.path.join(d, f)
                backups.append({"filename": f, "size": os.path.getsize(path), "date": f})
        self.send_json_response(backups)

    def download_backup(self, filename):
        if not filename or not filename.endswith('.db'):
            self.send_error(400, "Bad Request")
            return
        backup_path = os.path.join(self._backup_dir(), filename)
        if not os.path.exists(backup_path):
            self.send_error(404, "File Not Found")
            return
        try:
            with open(backup_path, 'rb') as f:
                data = f.read()
            self.send_response(200)
            self.send_header('Content-Type', 'application/octet-stream')
            self.send_header('Content-Disposition', 'attachment; filename="' + filename + '"')
            self.send_header('Content-Length', str(len(data)))
            self.send_header('Cache-Control', 'no-store')
            self.end_headers()
            self.wfile.write(data)
        except Exception:
            self.send_error(500, "Download failed")

    def create_backup(self):
        import time
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_name = f"backup_{timestamp}.db"
        src = db.DATABASE_PATH
        dst = os.path.join(self._backup_dir(), backup_name)
        try:
            shutil.copy2(src, dst)
            self.send_json_response({"success": True, "filename": backup_name})
        except Exception as e:
            self.send_json_response({"error": f"فشل في إنشاء النسخة الاحتياطية: {str(e)}"}, 500)

    def restore_backup(self, data):
        filename = data.get('filename', '')
        if not filename:
            self.send_json_response({"error": "اسم الملف مطلوب"}, 400)
            return
        backup_path = os.path.join(self._backup_dir(), filename)
        if not os.path.exists(backup_path) or not filename.endswith('.db'):
            self.send_json_response({"error": "ملف النسخة الاحتياطية غير موجود"}, 404)
            return
        try:
            shutil.copy2(backup_path, db.DATABASE_PATH)
            self.send_json_response({"success": True})
        except Exception as e:
            self.send_json_response({"error": f"فشل في الاستعادة: {str(e)}"}, 500)

    def delete_backup(self, data):
        filename = data.get('filename', '')
        if not filename:
            self.send_json_response({"error": "اسم الملف مطلوب"}, 400)
            return
        backup_path = os.path.join(self._backup_dir(), filename)
        if os.path.exists(backup_path):
            os.remove(backup_path)
        self.send_json_response({"success": True})

    # ==================== INVENTORY XLSX EXPORT ====================
    def export_inventory_xlsx(self):
        try:
            import openpyxl
            import openpyxl.styles
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "BVC Stock"
            ws.sheet_view.rightToLeft = True
            headers = ['كود الصنف', 'اسم الصنف', 'الكمية', 'الوحدة', 'سعر القطاعي', 'اخر شراء', 'متوسط الشراء', 'سعر السوق', 'حد الطلب', 'الوصف']
            header_font = openpyxl.styles.Font(bold=True, size=11)
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            rows = db.query_db("SELECT * FROM inventory ORDER BY item_name ASC")
            for r_idx, row in enumerate(rows, 2):
                ws.cell(row=r_idx, column=1, value=row['item_code'])
                ws.cell(row=r_idx, column=2, value=row['item_name'])
                ws.cell(row=r_idx, column=3, value=row['quantity'])
                ws.cell(row=r_idx, column=4, value=row['unit'])
                ws.cell(row=r_idx, column=5, value=row['wholesale_price'] or 0)
                ws.cell(row=r_idx, column=6, value=row['last_purchase_price'] or 0)
                ws.cell(row=r_idx, column=7, value=row['avg_purchase_price'] or 0)
                ws.cell(row=r_idx, column=8, value=row['market_price'] or 0)
                ws.cell(row=r_idx, column=9, value=row['min_stock'])
                ws.cell(row=r_idx, column=10, value=row['description'] or '')
            for col_letter in ['A','B','C','D','E','F','G','H','I','J']:
                ws.column_dimensions[col_letter].width = 18
            import io
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            xlsx_bytes = buf.read()
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', 'attachment; filename="BVC_Stock.xlsx"')
            self.send_header('Content-Length', str(len(xlsx_bytes)))
            self.send_header('Cache-Control', 'no-store')
            self.end_headers()
            self.wfile.write(xlsx_bytes)
        except Exception as e:
            self.send_json_response({"error": f"فشل التصدير: {str(e)}"}, 500)

    # ==================== INVENTORY XLSX IMPORT (BATCH) ====================
    _import_rows = {}
    _import_counter = 0

    def import_inventory_xlsx(self):
        parsed = self.parse_multipart()
        if not parsed or not parsed.get('filedata'):
            self.send_json_response({"error": "لم يتم رفع ملف"}, 400)
            return
        filename = parsed.get('filename', '')
        filedata = parsed['filedata']
        if not filename or not (filename.endswith('.xlsx') or filename.endswith('.xls')):
            self.send_json_response({"error": "يجب أن يكون الملف بصيغة xlsx فقط"}, 400)
            return
        try:
            import openpyxl
            import io
            buf = io.BytesIO(filedata)
            wb = openpyxl.load_workbook(buf, read_only=True)
            ws = wb.active
            raw_rows = list(ws.iter_rows(min_row=2, values_only=True))
            rows = []
            for row in raw_rows:
                if not row or not row[0]:
                    continue
                rows.append({
                    'item_code': str(row[0]).strip(),
                    'item_name': str(row[1]).strip() if row[1] else str(row[0]).strip(),
                    'quantity': int(row[2]) if row[2] is not None else 0,
                    'unit': str(row[3]).strip() if row[3] else 'قطعة',
                    'wholesale_price': float(row[4]) if row[4] is not None else 0,
                    'last_purchase_price': float(row[5]) if row[5] is not None else 0,
                    'avg_purchase_price': float(row[6]) if row[6] is not None else 0,
                    'market_price': float(row[7]) if row[7] is not None else 0,
                    'min_stock': int(row[8]) if row[8] is not None else 5,
                    'description': str(row[9]).strip() if row[9] else '',
                })
            BVCRequestHandler._import_counter += 1
            upload_id = str(BVCRequestHandler._import_counter)
            BVCRequestHandler._import_rows[upload_id] = rows
            self.send_json_response({"success": True, "upload_id": upload_id, "total": len(rows)})
        except Exception as e:
            self.send_json_response({"error": f"فشل قراءة الملف: {str(e)}"}, 500)

    def import_process_batch(self, data):
        upload_id = str(data.get('upload_id', ''))
        batch_size = int(data.get('batch_size', 50))
        if not upload_id or upload_id not in BVCRequestHandler._import_rows:
            self.send_json_response({"error": "معرّف الرفع غير صالح أو انتهت صلاحيته"}, 400)
            return
        rows = BVCRequestHandler._import_rows[upload_id]
        total = len(rows)
        processed = int(data.get('processed', 0))
        batch_added = 0
        batch_updated = 0
        batch_end = min(processed + batch_size, total)
        for i in range(processed, batch_end):
            r = rows[i]
            existing = db.query_db("SELECT id FROM inventory WHERE item_code=?", (r['item_code'],), one=True)
            if existing:
                db.execute_db('''
                    UPDATE inventory SET item_name=?, quantity=?, unit=?, wholesale_price=?, last_purchase_price=?, avg_purchase_price=?, market_price=?, min_stock=?, description=?, updated_at=CURRENT_TIMESTAMP
                    WHERE item_code=?
                ''', (r['item_name'], r['quantity'], r['unit'], r['wholesale_price'], r['last_purchase_price'], r['avg_purchase_price'], r['market_price'], r['min_stock'], r['description'], r['item_code']))
                batch_updated += 1
            else:
                db.execute_db('''
                    INSERT INTO inventory (item_name, item_code, quantity, unit, wholesale_price, last_purchase_price, avg_purchase_price, market_price, min_stock, description)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (r['item_name'], r['item_code'], r['quantity'], r['unit'], r['wholesale_price'], r['last_purchase_price'], r['avg_purchase_price'], r['market_price'], r['min_stock'], r['description']))
                batch_added += 1
        new_processed = batch_end
        done = new_processed >= total
        if done:
            del BVCRequestHandler._import_rows[upload_id]
        self.send_json_response({
            "success": True,
            "processed": new_processed,
            "total": total,
            "batch_added": batch_added,
            "batch_updated": batch_updated,
            "done": done,
        })

    def upload_backup(self):
        parsed = self.parse_multipart()
        if not parsed or not parsed.get('filedata'):
            self.send_json_response({"error": "لم يتم رفع ملف. يرجى اختيار ملف .db"}, 400)
            return
        filename = parsed.get('filename', '')
        filedata = parsed['filedata']
        should_restore = parsed.get('restore', 'false') == 'true'
        if not filename or not filename.endswith('.db'):
            self.send_json_response({"error": "يجب أن يكون الملف بصيغة .db فقط"}, 400)
            return
        safe_name = os.path.basename(filename)
        d = self._backup_dir()
        save_path = os.path.join(d, safe_name)
        try:
            with open(save_path, 'wb') as f:
                f.write(filedata)
        except Exception as e:
            self.send_json_response({"error": f"فشل في حفظ الملف: {str(e)}"}, 500)
            return
        if should_restore:
            try:
                shutil.copy2(save_path, db.DATABASE_PATH)
                self.send_json_response({"success": True, "filename": safe_name, "restored": True})
            except Exception as e:
                self.send_json_response({"error": f"تم الحفظ لكن فشلت الاستعادة: {str(e)}"}, 500)
        else:
            self.send_json_response({"success": True, "filename": safe_name, "restored": False})


    # ==================== GENERIC DATA EXPORT ====================
    def export_data(self, data):
        fmt = data.get('format', 'csv')
        columns = data.get('columns', [])
        rows = data.get('rows', [])
        title = data.get('title', 'تصدير')
        filename = data.get('filename', 'export')

        if not columns or not rows:
            self.send_json_response({"error": "لا توجد بيانات للتصدير"}, 400)
            return

        try:
            if fmt == 'csv':
                self._export_csv(columns, rows, filename)
            elif fmt == 'xlsx':
                self._export_xlsx(columns, rows, title, filename)
            elif fmt == 'pdf':
                self._export_pdf(columns, rows, title, filename)
            else:
                self.send_json_response({"error": f"صيغة غير مدعومة: {fmt}"}, 400)
        except Exception as e:
            server_log(f"Export failed: {e}")
            self.send_json_response({"error": f"فشل التصدير: {str(e)}"}, 500)

    def _export_csv(self, columns, rows, filename):
        import csv, io
        output = io.StringIO()
        output.write('\ufeff')
        writer = csv.writer(output)
        writer.writerow([c['label'] for c in columns])
        for row in rows:
            writer.writerow([row.get(c['key'], '') for c in columns])
        content = output.getvalue().encode('utf-8-sig')
        self.send_response(200)
        self.send_header('Content-Type', 'text/csv; charset=utf-8-sig')
        self.send_header('Content-Disposition', f'attachment; filename="{filename}.csv"')
        self.send_header('Content-Length', str(len(content)))
        self.send_header('Cache-Control', 'no-store')
        self.end_headers()
        self.wfile.write(content)

    def _export_xlsx(self, columns, rows, title, filename):
        import openpyxl
        import openpyxl.styles
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]
        ws.sheet_view.rightToLeft = True
        header_font = openpyxl.styles.Font(bold=True, size=11)
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col['label'])
            cell.font = header_font
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        for r_idx, row in enumerate(rows, 2):
            for c_idx, col in enumerate(columns, 1):
                ws.cell(row=r_idx, column=c_idx, value=row.get(col['key'], ''))
        for ci in range(1, len(columns) + 1):
            ws.column_dimensions[chr(64 + ci) if ci <= 26 else 'A'].width = 18
        import io
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        xlsx_bytes = buf.read()
        self.send_response(200)
        self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.send_header('Content-Disposition', f'attachment; filename="{filename}.xlsx"')
        self.send_header('Content-Length', str(len(xlsx_bytes)))
        self.send_header('Cache-Control', 'no-store')
        self.end_headers()
        self.wfile.write(xlsx_bytes)

    def _export_pdf(self, columns, rows, title, filename):
        html = self._render_export_html(columns, rows, title)
        content = html.encode('utf-8')
        self.send_response(200)
        self.send_header('Content-Type', 'text/html; charset=utf-8')
        self.send_header('Content-Disposition', f'inline; filename="{filename}.html"')
        self.send_header('Content-Length', str(len(content)))
        self.send_header('Cache-Control', 'no-store')
        self.end_headers()
        self.wfile.write(content)

    def _render_export_html(self, columns, rows, title):
        thead = '<tr>' + ''.join(f'<th>{c["label"]}</th>' for c in columns) + '</tr>'
        trows = []
        for row in rows:
            cells = ''.join(f'<td>{str(row.get(c["key"], "") or "")}</td>' for c in columns)
            trows.append(f'<tr>{cells}</tr>')
        tbody = '\n'.join(trows)
        return f'''<!DOCTYPE html><html dir="rtl"><head><meta charset="utf-8">
<style>
  @media print {{ body {{ padding:0;margin:0; }} }}
  body {{ font-family:"Tajawal","Segoe UI",sans-serif; padding:20px; }}
  h1 {{ font-size:1.2rem; text-align:center; margin-bottom:16px; color:#1e293b; }}
  table {{ width:100%; border-collapse:collapse; font-size:0.8rem; }}
  th,td {{ border:1px solid #ddd; padding:6px 8px; text-align:center; }}
  th {{ background:#0d9488; color:#fff; }}
  tr:nth-child(even) {{ background:#f8fafc; }}
  .print-btn {{ display:block; margin:16px auto; padding:8px 24px; background:#0d9488; color:#fff; border:none; border-radius:6px; cursor:pointer; font-size:1rem; }}
  @media print {{ .print-btn {{ display:none; }} }}
</style></head><body>
<h1>{title}</h1>
<button class="print-btn" onclick="window.print()">🖨️ طباعة / حفظ PDF</button>
<table><thead>{thead}</thead><tbody>{tbody}</tbody></table>
<script>window.onload=function(){{setTimeout(function(){{window.print()}},500)}}</script>
</body></html>'''


def run_server():
    try:
        os.remove(LOG_FILE)
    except Exception:
        pass
    server_log("Server starting...")
    server_log(f"BASE_DIR: {BASE_DIR}")
    server_log(f"PUBLIC_DIR: {PUBLIC_DIR} exists={os.path.exists(PUBLIC_DIR)}")
    server_log(f"DB path: {db.DATABASE_PATH}")

    try:
        db.init_db()
        server_log("init_db() completed successfully")
    except Exception as e:
        server_log(f"init_db() FAILED: {e}")
        import traceback
        server_log(traceback.format_exc())

    # Use ThreadingHTTPServer if available (Python 3.7+), else fallback to ThreadingTCPServer
    try:
        class ThreadingHTTPServer(socketserver.ThreadingMixIn, http.server.HTTPServer):
            pass
        server_address = ('', PORT)
        httpd = ThreadingHTTPServer(server_address, BVCRequestHandler)
        server_log("ThreadingHTTPServer bound to port " + str(PORT))
    except Exception as e:
        server_log(f"ThreadingHTTPServer failed: {e}, using basic HTTPServer")
        server_address = ('', PORT)
        httpd = http.server.HTTPServer(server_address, BVCRequestHandler)

    print(f"BVC System Server running on http://localhost:{PORT}")
    server_log("Server ready and serving")
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\nStopping BVC System Server...")
        httpd.server_close()

if __name__ == '__main__':
    run_server()
