# -*- coding: utf-8 -*-
import os
import sys
import io
import re

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

os.chdir(os.path.dirname(os.path.abspath(__file__)))

from db import init_db, get_db_connection

def clean_name(name):
    if not name:
        return ''
    name = str(name).strip()
    name = re.sub(r'[ـــ]+', '', name)  # remove tashkeel elongation
    name = re.sub(r'\s+', ' ', name)
    name = name.strip()
    return name

def to_float(s):
    if s is None:
        return 0.0
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0

def to_int(s):
    if s is None:
        return 0
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 0

def import_employees(wb):
    """Import employees from sheet."""
    print("\n=== Importing Employees ===")
    ws = wb['الراتب الاسبوعي']
    conn = get_db_connection()
    cursor = conn.cursor()
    count = 0
    
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3:  # skip headers
            continue
        vals = list(row)
        name = clean_name(vals[1]) if len(vals) > 1 else ''
        weekly_rate = to_float(vals[2]) if len(vals) > 2 else 0
        
        if not name or not weekly_rate:
            continue
        
        try:
            cursor.execute('''
                INSERT OR IGNORE INTO employees (name, employee_code, phone, national_id, pay_type, rate, default_shift, status)
                VALUES (?, '', '', '', 'shift', ?, '1', 'active')
            ''', (name, weekly_rate))
            if cursor.rowcount > 0:
                count += 1
        except Exception as e:
            print(f"  Error inserting employee {name}: {e}")
    
    conn.commit()
    conn.close()
    print(f"Imported {count} employees")
    return count

def import_debts(wb):
    """Import debts from sheet."""
    print("\n=== Importing Debts ===")
    ws = wb['مديونيات']
    conn = get_db_connection()
    cursor = conn.cursor()
    count = 0
    
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3:
            continue
        vals = list(row)
        name = clean_name(vals[2]) if len(vals) > 2 else ''
        debt = to_float(vals[3]) if len(vals) > 3 else 0
        
        if not name or not debt:
            continue
        
        # Find employee by name
        emp = cursor.execute("SELECT id FROM employees WHERE name LIKE ?", (f'%{name}%',)).fetchone()
        if not emp:
            continue
        
        emp_id = emp[0]
        try:
            cursor.execute('''
                INSERT INTO employee_ledger (employee_id, date, type, amount, description)
                VALUES (?, '2026-07-01', 'debt_add', ?, 'مديونية من النظام القديم')
            ''', (emp_id, -debt))
            count += 1
        except Exception as e:
            print(f"  Error inserting debt for {name}: {e}")
    
    conn.commit()
    conn.close()
    print(f"Imported {count} debt records")
    return count

def import_payroll(wb):
    """Import payroll data from القبض الاسبوعي sheet."""
    print("\n=== Importing Payroll ===")
    ws = wb['القبض الاسبوعي']
    conn = get_db_connection()
    cursor = conn.cursor()
    count = 0
    
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 4:  # skip headers
            continue
        vals = list(row)
        name = clean_name(vals[2]) if len(vals) > 2 else ''
        if not name:
            continue
        
        attend_pay = to_float(vals[3]) if len(vals) > 3 else 0
        absent_pay = to_float(vals[4]) if len(vals) > 4 else 0
        late = to_float(vals[5]) if len(vals) > 5 else 0
        ot = to_float(vals[6]) if len(vals) > 6 else 0
        withdraw = to_float(vals[7]) if len(vals) > 7 else 0
        deduct = to_float(vals[8]) if len(vals) > 8 else 0
        incentive = to_float(vals[9]) if len(vals) > 9 else 0
        debt_remaining = to_float(vals[11]) if len(vals) > 11 else 0
        days_worked = to_float(vals[14]) if len(vals) > 14 else 0
        
        emp = cursor.execute("SELECT id FROM employees WHERE name LIKE ?", (f'%{name}%',)).fetchone()
        if not emp:
            continue
        
        emp_id = emp[0]
        net = attend_pay + ot - absent_pay - late - withdraw - deduct
        
        try:
            # Create payroll record
            cursor.execute('''
                INSERT INTO payroll (employee_id, start_date, end_date, total_hours, total_shifts, base_salary, total_bonuses, total_deductions, total_loans_deducted, net_salary)
                VALUES (?, '2026-07-01', '2026-07-17', ?, ?, ?, ?, ?, ?, ?)
            ''', (emp_id, days_worked * 8, int(days_worked), attend_pay, incentive, late + deduct, withdraw, max(0, net)))
            payroll_id = cursor.lastrowid
            
            # Add ledger entries
            if attend_pay > 0:
                cursor.execute('''
                    INSERT INTO employee_ledger (employee_id, date, type, amount, description, payroll_id)
                    VALUES (?, '2026-07-17', 'salary_earned', ?, 'راتب حضور', ?)
                ''', (emp_id, attend_pay, payroll_id))
            if ot > 0:
                cursor.execute('''
                    INSERT INTO employee_ledger (employee_id, date, type, amount, description, payroll_id)
                    VALUES (?, '2026-07-17', 'bonus', ?, 'اضافي', ?)
                ''', (emp_id, ot, payroll_id))
            if incentive > 0:
                cursor.execute('''
                    INSERT INTO employee_ledger (employee_id, date, type, amount, description, payroll_id)
                    VALUES (?, '2026-07-17', 'monthly_incentive', ?, 'حافز', ?)
                ''', (emp_id, incentive, payroll_id))
            if withdraw > 0:
                cursor.execute('''
                    INSERT INTO employee_ledger (employee_id, date, type, amount, description, payroll_id)
                    VALUES (?, '2026-07-17', 'withdrawal', ?, 'مسحوبات', ?)
                ''', (emp_id, -withdraw, payroll_id))
            if late > 0:
                cursor.execute('''
                    INSERT INTO employee_ledger (employee_id, date, type, amount, description, payroll_id)
                    VALUES (?, '2026-07-17', 'deduction', ?, 'تاخير', ?)
                ''', (emp_id, -late, payroll_id))
            if deduct > 0:
                cursor.execute('''
                    INSERT INTO employee_ledger (employee_id, date, type, amount, description, payroll_id)
                    VALUES (?, '2026-07-17', 'deduction', ?, 'خصم', ?)
                ''', (emp_id, -deduct, payroll_id))
            
            count += 1
        except Exception as e:
            print(f"  Error for {name}: {e}")
    
    conn.commit()
    conn.close()
    print(f"Imported {count} payroll records")
    return count

def import_expenses(wb):
    """Import expenses and treasury transactions from المصروفات sheet."""
    print("\n=== Importing Expenses & Treasury ===")
    ws = wb['المصروفات']
    conn = get_db_connection()
    cursor = conn.cursor()
    expense_count = 0
    treasury_count = 0
    balance = 0
    
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 8:
            continue
        vals = list(row)
        
        date_raw = vals[1]
        incoming = to_float(vals[3])
        outgoing = to_float(vals[4])
        desc = str(vals[5]).strip() if vals[5] else ''
        note = str(vals[6]).strip() if vals[6] else ''
        
        if not date_raw or date_raw in ['******', '********', '************']:
            continue
        
        if hasattr(date_raw, 'strftime'):
            date_str = date_raw.strftime('%Y-%m-%d')
        else:
            date_str = str(date_raw)[:10]
        
        if not incoming and not outgoing:
            continue
        
        description = desc or note or ''
        
        if incoming > 0:
            # Deposit into treasury
            cursor.execute('''
                INSERT INTO cashbox (date, type, amount, source, description, category, balance_after)
                VALUES (?, 'deposit', ?, ?, ?, 'وارد', ?)
            ''', (date_str, incoming, description, description, balance + incoming))
            balance += incoming
            treasury_count += 1
            
            if 'ابورحمه' in description:
                pass  # Owner deposit
        
        if outgoing > 0:
            # Withdrawal from treasury
            cursor.execute('''
                INSERT INTO cashbox (date, type, amount, source, description, category, balance_after)
                VALUES (?, 'withdrawal', ?, ?, ?, 'مصروفات', ?)
            ''', (date_str, outgoing, description, description, balance - outgoing))
            balance -= outgoing
            treasury_count += 1
            
            # Also record as factory expense
            category = 'مصروفات عامة'
            if 'بенزين' in desc or 'وقود' in desc or 'بنزين' in desc:
                category = 'وقود'
            elif 'رواتب' in desc or 'مسحوبات' in desc:
                category = 'رواتب'
            elif 'جماعيه' in desc or 'جمعيه' in desc:
                category = 'جمعيه'
            elif 'دليفري' in desc or 'مواصلات' in desc:
                category = 'مواصلات'
            elif 'عشا' in desc or 'غدا' in desc or 'اكل' in desc or 'سهرات' in desc:
                category = 'طعام'
            elif 'علاج' in desc:
                category = 'علاج'
            
            cursor.execute('''
                INSERT INTO factory_expenses (date, amount, category, description)
                VALUES (?, ?, ?, ?)
            ''', (date_str, outgoing, category, description))
            expense_count += 1
    
    conn.commit()
    conn.close()
    print(f"Imported {treasury_count} treasury transactions, {expense_count} factory expenses")
    print(f"Final balance: {balance}")
    return treasury_count, expense_count

def import_inventory():
    """Import inventory items from الاصناف.xls (XML Spreadsheet format)."""
    print("\n=== Importing Inventory ===")
    import xml.etree.ElementTree as ET
    ns = {'ss': 'urn:schemas-microsoft-com:office:spreadsheet'}
    
    xls_files = [f for f in os.listdir('.') if f.endswith('.xls')]
    if not xls_files:
        print("No .xls file found!")
        return 0
    
    try:
        tree = ET.parse(xls_files[0])
    except Exception as e:
        print(f"Failed to parse XML: {e}")
        return 0
    
    root = tree.getroot()
    conn = get_db_connection()
    cursor = conn.cursor()
    count = 0
    
    for sheet in root.findall('.//ss:Worksheet', ns):
        for row in sheet.findall('.//ss:Row', ns):
            grid = {}
            col = 1
            for cell in row.findall('ss:Cell', ns):
                explicit_idx = cell.attrib.get('{urn:schemas-microsoft-com:office:spreadsheet}Index')
                if explicit_idx:
                    col = int(explicit_idx)
                merge = cell.attrib.get('{urn:schemas-microsoft-com:office:spreadsheet}MergeAcross', '0')
                merge_across = int(merge) if merge else 0
                data_el = cell.find('ss:Data', ns)
                val = data_el.text.strip() if data_el is not None and data_el.text else ''
                grid[col] = val
                col += 1 + merge_across
            
            item_code = grid.get(32, '').strip()
            item_name = grid.get(30, '').strip()
            unit = grid.get(28, '').strip()
            wholesale = grid.get(17, '')
            last_p = grid.get(8, '')
            avg_p = grid.get(11, '')
            market = grid.get(1, '')
            min_retail = grid.get(15, '')
            
            if not item_name or not item_code or item_code == 'رقم الصنف':
                continue
            
            wholesale_f = to_float(wholesale)
            last_p_f = to_float(last_p)
            avg_p_f = to_float(avg_p)
            market_f = to_float(market)
            min_retail_f = to_float(min_retail)
            
            try:
                cursor.execute('''
                    INSERT OR IGNORE INTO inventory (item_name, item_code, quantity, unit, purchase_price, sale_price, min_stock, description, wholesale_price, market_price, last_purchase_price, avg_purchase_price, min_retail_price)
                    VALUES (?, ?, 0, ?, ?, ?, 5, '', ?, ?, ?, ?, ?)
                ''', (item_name, item_code, unit, last_p_f, wholesale_f, wholesale_f, market_f, last_p_f, avg_p_f, min_retail_f))
                if cursor.rowcount > 0:
                    count += 1
            except Exception as e:
                print(f"  Error inserting item {item_code}: {e}")
    
    conn.commit()
    conn.close()
    print(f"Imported {count} inventory items")
    return count

def import_items_data(wb):
    """Import/update items from xlsm data sheet."""
    print("\n=== Importing Items Data (xlsm) ===")
    ws = wb['data']
    conn = get_db_connection()
    cursor = conn.cursor()
    count = 0
    
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 1:  # skip header
            continue
        vals = list(row)
        code = str(vals[1]).strip() if len(vals) > 1 and vals[1] else ''
        name = str(vals[2]).strip() if len(vals) > 2 and vals[2] else ''
        price = to_float(vals[3]) if len(vals) > 3 else 0
        
        if not code or not name:
            continue
        
        if code.isdigit() and int(code) > 1000000:
            continue  # skip row numbers
        
        try:
            # Try to update existing item first
            existing = cursor.execute("SELECT id FROM inventory WHERE item_code = ?", (code,)).fetchone()
            if existing:
                if price > 0:
                    cursor.execute("UPDATE inventory SET sale_price = ?, wholesale_price = ? WHERE item_code = ?", (price, price, code))
            else:
                cursor.execute('''
                    INSERT OR IGNORE INTO inventory (item_name, item_code, quantity, unit, purchase_price, sale_price, min_stock, description)
                    VALUES (?, ?, 0, 'قطعة', ?, ?, 5, '')
                ''', (name, code, price, price))
                if cursor.rowcount > 0:
                    count += 1
        except Exception as e:
            pass
    
    conn.commit()
    conn.close()
    print(f"Imported {count} new items from data sheet")
    return count

if __name__ == '__main__':
    print("=" * 60)
    print("BVC Al-Rahma Data Import")
    print("=" * 60)
    
    # 1. Initialize DB
    init_db()
    
    # 2. Find and load xlsm file
    xlsm_files = [f for f in os.listdir('.') if f.endswith('.xlsm')]
    if not xlsm_files:
        print("No xlsm file found!")
        sys.exit(1)
    
    import openpyxl
    wb = openpyxl.load_workbook(xlsm_files[0], read_only=True, keep_vba=True, data_only=True)
    
    # 3. Import all data
    import_employees(wb)
    import_debts(wb)
    import_payroll(wb)
    import_expenses(wb)
    import_items_data(wb)
    wb.close()
    
    # 4. Import inventory from XML xls
    import_inventory()
    
    # 5. Summary
    conn = get_db_connection()
    emp_count = conn.execute("SELECT COUNT(*) FROM employees").fetchone()[0]
    inv_count = conn.execute("SELECT COUNT(*) FROM inventory").fetchone()[0]
    cash_count = conn.execute("SELECT COUNT(*) FROM cashbox").fetchone()[0]
    ledger_count = conn.execute("SELECT COUNT(*) FROM employee_ledger").fetchone()[0]
    balance = conn.execute("SELECT COALESCE(SUM(CASE WHEN type='deposit' THEN amount ELSE -amount END), 0) FROM cashbox").fetchone()[0]
    conn.close()
    
    print("\n" + "=" * 60)
    print("Import Summary:")
    print(f"  Employees: {emp_count}")
    print(f"  Inventory Items: {inv_count}")
    print(f"  Treasury Transactions: {cash_count}")
    print(f"  Employee Ledger Entries: {ledger_count}")
    print(f"  Treasury Balance: {balance:,.2f} EGP")
    print("=" * 60)
